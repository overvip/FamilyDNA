#! python3
# -*- coding:utf-8 -*-
from tkinter import *

import openpyxl


# from openpyxl.utils import column_index_from_string


def rows_merged(i):
    """计算每个家系人数，人数为函数返回值+1。"""
    m_list = wsVillage.merged_cells
    for cell in m_list:
        if cell.max_col == cell.min_col == 2:
            if cell.min_row <= i <= cell.max_row:
                return cell.max_row - cell.min_row
    return 0


def get_member_string(row, r):
    member_string = ':{'
    for i in range(r +1):
        member_string += wsVillage.cell(row + i, 3).value + ' '
    member_string = member_string.strip() + '}\n'
    return member_string


root = Tk()
root.title('采血统计')
tb = Text(root, width=80, height=40)
sl = Scrollbar(root, command=tb.yview)
tb.configure(yscrollcommand=sl.set)
tb.tag_configure('color', foreground='blue',
                 font=('Tempus Sans ITC', 14, 'bold'))

f = 'C:/Users/shine/onedrive/社区采血名单.xlsx'
villages = ('鼎美村', '后柯村', '芸美村', '凤山社区', '东瑶村', '贞岱村')
wb = openpyxl.load_workbook(f)
for village in villages:
    tot_families = coll_families = tot_persons = coll_persons = 0
    complete_fam_persons = 0
    if village in wb.sheetnames:
        wsVillage = wb[village]

        #   从cell[b3]开始判断海家系，若该行b列非合并单元，则为单人家系，
        #   否则为合并单元，该家系需采集多于1人，继续判断该合并单元占几行，
        #   即该家系要采集几名成员。再检查j列标记是否采集。
        str_stat_info = '已完成家系：\n'
        row = 3
        while len(str(wsVillage.cell(row, 4).value).strip()) == 18:
            tot_families += 1
            r = rows_merged(row)
            if r > 0:
                flag_completion = True
                for j in range(r + 1):
                    if wsVillage.cell(row + j, 7).value is not None:
                        coll_persons += 1
                    else:
                        flag_completion = False
                    tot_persons += 1
                if flag_completion:
                    coll_families += 1
                    complete_fam_persons += r+1
                    str_stat_info += wsVillage.cell(row, 2).value + \
                                     get_member_string(row, r)
                    # tb.insert(END, wsVillage.cell(row,2).value+'\n', 'color')
                row = row + r + 1
            else:
                # tot_families += 1
                tot_persons += 1
                if wsVillage.cell(row, 7).value is not None:
                    coll_persons += 1
                    coll_families += 1
                    complete_fam_persons += 1
                    str_stat_info += wsVillage.cell(row, 2).value + \
                                     ':{' + wsVillage.cell(row, 3).value + \
                                     '}\n'
                    # tb.insert(END, wsVillage.cell(row,2).value+'\n', 'color')
                row += 1

        msgResult = village + '家系总数' + str(tot_families) + '个,已采集' + str(coll_families) + '个（共' + str(complete_fam_persons) + '人)，采集总数' + str(tot_persons) + '个，已采集' +  str(coll_persons) + '个\n\n'
        # print(msgResult)
        tb.insert(END, msgResult, 'color')
        tb.insert(END, str_stat_info + '\n', 'color')
        # input('按回车显示下一个村')
        # wsVillage['a1'].value = msgResult
# print('统计完毕')
# wb.save(f)

tb.config(state=DISABLED)
tb.pack(side=LEFT)
sl.pack(side=RIGHT, fill=Y)
root.mainloop()
